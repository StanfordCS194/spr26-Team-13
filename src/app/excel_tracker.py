"""Excel-based workout program tracker.

Reads prescribed exercises from the Jeff Nippard spreadsheet layout and writes
actual loads back into the same file. Column layout expected:
  B = week/day label, C = exercise, D = warmup sets, E = working sets,
  F = reps, G = load (prescribed or blank for accessories), H = %1RM,
  I = RPE, J = rest, K = notes, M = user comment, N = date logged
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore[assignment]


@dataclass
class ExerciseRow:
    row_number: int
    week: int
    day: int
    day_label: str
    exercise_name: str
    warmup_sets: int | None
    working_sets: int | None
    prescribed_reps: str | None
    prescribed_load: str | None
    percent_1rm: str | None
    rpe: str | None
    rest: str | None
    notes: str | None
    comment: str | None
    logged_date: str | None


_COL_B = 1   # day/week label (0-indexed)
_COL_C = 2   # exercise name
_COL_D = 3   # warmup sets
_COL_E = 4   # working sets
_COL_F = 5   # reps
_COL_G = 6   # load
_COL_H = 7   # %1RM
_COL_I = 8   # RPE
_COL_J = 9   # rest
_COL_K = 10  # notes
_COL_M = 12  # user comment
_COL_N = 13  # date logged


def _to_str(val: Any) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return None  # openpyxl misparses "7-8" RPE ranges as dates
    s = str(val).strip()
    return s or None


def _to_int(val: Any) -> int | None:
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


def _parse_rpe(val: Any) -> str | None:
    if isinstance(val, datetime):
        return None  # Excel date misparse of "8/7"-style RPE ranges
    if isinstance(val, (int, float)):
        return str(int(val)) if float(val).is_integer() else str(val)
    return _to_str(val)


class ExcelTracker:
    SHEET = "5x"

    def __init__(self, file_path: str | Path) -> None:
        if load_workbook is None:
            raise ImportError("openpyxl is required: pip install openpyxl")
        self.path = Path(file_path)
        self._rows: list[ExerciseRow] = []
        self._load()

    def _load(self) -> None:
        wb = load_workbook(self.path, data_only=True)
        ws = wb[self.SHEET]
        current_week = 0
        current_day = 0
        current_day_label = ""

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            b = row[_COL_B] if len(row) > _COL_B else None
            c = row[_COL_C] if len(row) > _COL_C else None

            if isinstance(b, str):
                if re.match(r"week\s+\d+", b, re.IGNORECASE):
                    m = re.search(r"\d+", b)
                    if m:
                        current_week = int(m.group())
                        current_day = 0
                    continue
                if "rest" in b.upper():
                    continue
                # New training day
                if b.strip():
                    current_day += 1
                    current_day_label = b.strip().rstrip(":")

            if not isinstance(c, str) or not c.strip():
                continue
            if current_week == 0:
                continue
            if c.strip().lower() == "exercise":
                continue

            g = row[_COL_G] if len(row) > _COL_G else None
            m_val = row[_COL_M] if len(row) > _COL_M else None
            n_val = row[_COL_N] if len(row) > _COL_N else None

            self._rows.append(ExerciseRow(
                row_number=row_idx,
                week=current_week,
                day=current_day,
                day_label=current_day_label,
                exercise_name=c.strip(),
                warmup_sets=_to_int(row[_COL_D] if len(row) > _COL_D else None),
                working_sets=_to_int(row[_COL_E] if len(row) > _COL_E else None),
                prescribed_reps=_to_str(row[_COL_F] if len(row) > _COL_F else None),
                prescribed_load=_to_str(g),
                percent_1rm=_to_str(row[_COL_H] if len(row) > _COL_H else None),
                rpe=_parse_rpe(row[_COL_I] if len(row) > _COL_I else None),
                rest=_to_str(row[_COL_J] if len(row) > _COL_J else None),
                notes=_to_str(row[_COL_K] if len(row) > _COL_K else None),
                comment=_to_str(m_val),
                logged_date=_to_str(n_val),
            ))

        wb.close()

    def get_workout(self, week: int, day: int) -> list[ExerciseRow]:
        """Return prescribed exercises for a specific week and day number (1-5)."""
        return [r for r in self._rows if r.week == week and r.day == day]

    def get_day_labels(self, week: int) -> list[str]:
        """Return distinct day labels for a week, e.g. ['FULL BODY 1', ...]."""
        seen: list[str] = []
        for r in self._rows:
            if r.week == week and r.day_label not in seen:
                seen.append(r.day_label)
        return seen

    def query_history(self, exercise_name: str) -> list[ExerciseRow]:
        """Return all rows across all weeks matching an exercise name."""
        needle = exercise_name.strip().lower()
        return [
            r for r in self._rows
            if needle in r.exercise_name.lower() or r.exercise_name.lower() in needle
        ]

    def log_set(
        self,
        week: int,
        day: int,
        exercise_name: str,
        actual_load: str | float | int,
        *,
        comment: str | None = None,
        log_date: date | None = None,
    ) -> bool:
        """
        Write completed set data back into the spreadsheet.

        - actual_load → column G only if the cell is currently blank
          (pre-calculated loads for main lifts are left untouched)
        - comment → column M (always written if provided)
        - log_date → column N (always written if provided)

        Returns True if a matching exercise row was found and saved.
        """
        needle = exercise_name.strip().lower()
        match = next(
            (r for r in self._rows if r.week == week and r.day == day and needle in r.exercise_name.lower()),
            None,
        )
        if match is None:
            return False

        wb = load_workbook(self.path)
        ws = wb[self.SHEET]
        row_idx = match.row_number

        g_cell = ws.cell(row=row_idx, column=_COL_G + 1)  # openpyxl is 1-indexed
        if g_cell.value is None:
            g_cell.value = actual_load

        if comment is not None:
            ws.cell(row=row_idx, column=_COL_M + 1).value = comment

        if log_date is not None:
            ws.cell(row=row_idx, column=_COL_N + 1).value = log_date.isoformat()

        wb.save(self.path)
        wb.close()

        match.prescribed_load = match.prescribed_load or str(actual_load)
        if comment:
            match.comment = comment
        if log_date:
            match.logged_date = log_date.isoformat()

        return True


def format_workout_summary(exercises: list[ExerciseRow]) -> str:
    """Format a workout's exercises into a human-readable coach reply."""
    if not exercises:
        return "No exercises found for that session."
    day_label = exercises[0].day_label
    lines = [f"{day_label}:"]
    for ex in exercises:
        sets = ex.working_sets or "?"
        reps = ex.prescribed_reps or "?"
        load = ex.prescribed_load
        load_str = f" at {load}" if load else ""
        lines.append(f"  {ex.exercise_name} — {sets}×{reps}{load_str}")
    return "\n".join(lines)


def format_history_summary(rows: list[ExerciseRow], exercise_name: str) -> str:
    """Format history rows into a coach-readable string."""
    filled = [r for r in rows if r.prescribed_load is not None]
    if not filled:
        return f"No logged loads found for {exercise_name} in your program."
    lines = [f"{exercise_name} history:"]
    for r in filled:
        date_str = f" ({r.logged_date})" if r.logged_date else ""
        comment_str = f" — {r.comment}" if r.comment else ""
        lines.append(f"  Week {r.week} {r.day_label}: {r.prescribed_load}{date_str}{comment_str}")
    return "\n".join(lines)
